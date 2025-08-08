#!/usr/bin/env python3
"""
Excel Housing Database Importer
Imports housing resources from Excel file into SQLite database
"""

import openpyxl
import logging
from pathlib import Path
from typing import Dict, List, Any
import sys
import os

# Add the app directory to Python path
sys.path.append('/app')

from housing.models import HousingResource, HousingDatabase

logger = logging.getLogger(__name__)

class HousingExcelImporter:
    """Imports housing data from Excel file to SQLite database"""
    
    def __init__(self, excel_path: str, db_path: str = "housing_resources.db"):
        self.excel_path = excel_path
        self.db_path = db_path
        self.database = HousingDatabase(db_path)
        
        # Column mapping from Excel to our model
        self.column_mapping = {
            'Facility Name': 'facility_name',
            'Physical Address': 'physical_address',
            'City': 'city',
            'State': 'state',
            'ZIP Code': 'zip_code',
            'Mailing Address': 'mailing_address',
            'Primary Phone': 'primary_phone',
            'Secondary Phone': 'secondary_phone',
            'Website URL': 'website_url',
            'Email Contact': 'email_contact',
            'Program Type': 'program_type',
            'Target Population': 'target_population',
            'Capacity': 'capacity',
            'Length of Stay': 'length_of_stay',
            'Hours of Operation': 'hours_of_operation',
            'Eligibility Criteria': 'eligibility_criteria',
            'Required Documentation': 'required_documentation',
            'Sobriety Requirements': 'sobriety_requirements',
            'Criminal Background Restrictions': 'criminal_background_restrictions',
            'Mental Health Requirements': 'mental_health_requirements',
            'Medical Requirements': 'medical_requirements',
            'Insurance Accepted': 'insurance_accepted',
            'Private Pay Options': 'private_pay_options',
            'Sliding Scale Fees': 'sliding_scale_fees',
            'Financial Assistance Programs': 'financial_assistance_programs',
            'Payment Plans Available': 'payment_plans_available',
            'Referral Requirements': 'referral_requirements',
            'Intake Process': 'intake_process',
            'Wait List Information': 'wait_list_information',
            'Contact Person': 'contact_person',
            'Clinical Services': 'clinical_services',
            'Life Skills Training': 'life_skills_training',
            'Job Placement Assistance': 'job_placement_assistance',
            'Transportation Services': 'transportation_services',
            'Medical Services': 'medical_services',
            'Additional Support Services': 'additional_support_services',
            'County': 'county',
            'Last Updated': 'last_updated'
        }
    
    def import_excel_data(self) -> Dict[str, Any]:
        """Import all housing data from Excel file"""
        if not os.path.exists(self.excel_path):
            raise FileNotFoundError(f"Excel file not found: {self.excel_path}")
        
        wb = openpyxl.load_workbook(self.excel_path)
        
        import_stats = {
            'total_imported': 0,
            'sheets_processed': 0,
            'errors': [],
            'by_sheet': {}
        }
        
        # Process each sheet except Summary
        for sheet_name in wb.sheetnames:
            if sheet_name == 'Summary':
                continue
                
            logger.info(f"Processing sheet: {sheet_name}")
            
            try:
                sheet_stats = self._process_sheet(wb[sheet_name], sheet_name)
                import_stats['by_sheet'][sheet_name] = sheet_stats
                import_stats['total_imported'] += sheet_stats['rows_imported']
                import_stats['sheets_processed'] += 1
                
            except Exception as e:
                error_msg = f"Error processing sheet {sheet_name}: {str(e)}"
                logger.error(error_msg)
                import_stats['errors'].append(error_msg)
        
        logger.info(f"Import completed: {import_stats['total_imported']} resources imported from {import_stats['sheets_processed']} sheets")
        return import_stats
    
    def _process_sheet(self, worksheet, sheet_name: str) -> Dict[str, Any]:
        """Process a single Excel sheet"""
        rows_imported = 0
        rows_with_errors = 0
        errors = []
        
        # Get header row
        headers = []
        for cell in worksheet[1]:
            headers.append(cell.value)
        
        # Process data rows
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), 2):
            if not row or not row[0]:  # Skip empty rows
                continue
            
            try:
                # Create resource data dictionary
                resource_data = {}
                
                for col_idx, header in enumerate(headers):
                    if col_idx < len(row) and header in self.column_mapping:
                        field_name = self.column_mapping[header]
                        value = row[col_idx]
                        
                        # Clean up the value
                        if value is not None:
                            value = str(value).strip()
                            if value.lower() in ['', 'n/a', 'none', 'null']:
                                value = ''
                        else:
                            value = ''
                        
                        resource_data[field_name] = value
                
                # Ensure required fields have values
                if not resource_data.get('facility_name'):
                    continue  # Skip rows without facility name
                
                # Set default values for missing fields
                resource_data.setdefault('state', 'CA')
                
                # Infer county from data or sheet
                if not resource_data.get('county'):
                    resource_data['county'] = self._infer_county_from_data(resource_data, sheet_name)
                
                resource_data.setdefault('program_type', self._infer_program_type_from_sheet(sheet_name))
                
                # Determine if facility is background-friendly
                resource_data['background_friendly'] = self._determine_background_friendly(
                    resource_data.get('criminal_background_restrictions', '')
                )
                
                # Create and save housing resource
                housing_resource = HousingResource(**resource_data)
                self.database.save_housing_resource(housing_resource)
                
                rows_imported += 1
                
            except Exception as e:
                error_msg = f"Error processing row {row_idx}: {str(e)}"
                logger.warning(error_msg)
                errors.append(error_msg)
                rows_with_errors += 1
        
        return {
            'rows_imported': rows_imported,
            'rows_with_errors': rows_with_errors,
            'errors': errors
        }
    
    def _infer_county_from_data(self, resource_data: Dict[str, str], sheet_name: str) -> str:
        """Infer county from resource data or default to Los Angeles"""
        
        # Check if county is already specified in the data
        if resource_data.get('county'):
            return resource_data['county']
        
        # Check city to determine county
        city = resource_data.get('city', '').lower()
        
        # San Diego County cities
        san_diego_cities = [
            'san diego', 'diego', 'north county san diego', 'east county san diego',
            'la jolla', 'del mar', 'encinitas', 'carlsbad', 'oceanside', 'vista',
            'escondido', 'poway', 'santee', 'el cajon', 'chula vista', 'imperial beach'
        ]
        
        # Check if city indicates San Diego County
        for sd_city in san_diego_cities:
            if sd_city in city:
                return "San Diego"
        
        # Los Angeles County cities (partial list)
        la_cities = [
            'los angeles', 'hollywood', 'beverly hills', 'santa monica', 'culver city',
            'pasadena', 'glendale', 'burbank', 'torrance', 'long beach', 'pomona',
            'west hollywood', 'manhattan beach', 'redondo beach', 'el segundo'
        ]
        
        # Check if city indicates Los Angeles County
        for la_city in la_cities:
            if la_city in city:
                return "Los Angeles"
        
        # Default to Los Angeles County (most facilities are there)
        return "Los Angeles"
    
    def _infer_program_type_from_sheet(self, sheet_name: str) -> str:
        """Infer program type from sheet name"""
        sheet_lower = sheet_name.lower()
        
        type_mapping = {
            'iop': 'IOP - Intensive Outpatient Program',
            'php': 'PHP - Partial Hospitalization Program',
            'sro': 'SRO - Single Room Occupancy',
            'mental health': 'Mental Health Housing',
            'sober living': 'Sober Living Housing',
            'transitional': 'Transitional Housing',
            'supportive': 'Supportive Housing',
            'halfway': 'Halfway House'
        }
        
        for key, value in type_mapping.items():
            if key in sheet_lower:
                return value
        
        return sheet_name  # Default to sheet name
    
    def _determine_background_friendly(self, background_restrictions: str) -> bool:
        """Determine if facility is background-friendly based on restrictions text"""
        if not background_restrictions:
            return True  # Default to background-friendly if no restrictions specified
        
        restrictions_lower = str(background_restrictions).lower()
        
        # Background-friendly indicators
        friendly_keywords = [
            'accepts individuals with criminal background',
            'case-by-case evaluation',
            'serves justice-involved',
            'specifically serves justice-involved individuals',
            'background-friendly'
        ]
        
        # Restrictive indicators
        restrictive_keywords = [
            'background check required',
            'background check may be required',
            'background check'
        ]
        
        # Check for background-friendly indicators first
        for keyword in friendly_keywords:
            if keyword in restrictions_lower:
                return True
        
        # Check for restrictive indicators
        for keyword in restrictive_keywords:
            if keyword in restrictions_lower:
                return False
        
        # Default to background-friendly if unclear
        return True
    
    def verify_import(self) -> Dict[str, Any]:
        """Verify the imported data"""
        stats = self.database.get_statistics()
        housing_types = self.database.get_all_housing_types()
        counties = self.database.get_all_counties()
        
        return {
            'database_stats': stats,
            'housing_types': housing_types,
            'counties': counties,
            'verification_passed': stats['total_resources'] > 0
        }


def main():
    """Main import function"""
    # Setup logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    # Excel file path
    excel_path = "/app/Housing Resource Database for Los Angeles and San Diego/Housing_Resources_Database.xlsx"
    
    # Database path
    db_path = "/app/housing_resources.db"
    
    try:
        # Create importer
        importer = HousingExcelImporter(excel_path, db_path)
        
        # Import data
        print("🏠 Starting housing database import...")
        import_stats = importer.import_excel_data()
        
        # Print results
        print(f"✅ Import completed!")
        print(f"📊 Total resources imported: {import_stats['total_imported']}")
        print(f"📑 Sheets processed: {import_stats['sheets_processed']}")
        
        if import_stats['errors']:
            print(f"⚠️  Errors encountered: {len(import_stats['errors'])}")
            for error in import_stats['errors'][:5]:  # Show first 5 errors
                print(f"   - {error}")
        
        # Verify import
        print("\n🔍 Verifying imported data...")
        verification = importer.verify_import()
        
        if verification['verification_passed']:
            print("✅ Verification passed!")
            print(f"📈 Database statistics:")
            stats = verification['database_stats']
            print(f"   - Total resources: {stats['total_resources']}")
            print(f"   - Background-friendly: {stats['background_friendly_count']} ({stats['background_friendly_percentage']:.1f}%)")
            print(f"   - Housing types: {len(verification['housing_types'])}")
            print(f"   - Counties: {len(verification['counties'])}")
        else:
            print("❌ Verification failed!")
        
        print(f"\n🗄️  Database created: {db_path}")
        
    except Exception as e:
        print(f"❌ Import failed: {e}")
        logger.error(f"Import failed: {e}", exc_info=True)
        return 1
    
    return 0


if __name__ == "__main__":
    exit(main())